from __future__ import annotations

import io
import os
import re
from typing import Any, Dict, List, Tuple

import requests
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

APP_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_XLSX = os.path.join(APP_DIR, "template.xlsx")

app = Flask(__name__)


# -----------------------------
# NoMoreAngel RAW parser
# -----------------------------
_KEYVAL_RE = re.compile(r"^\s*\[(.*?)\]\s*=>\s*(.*)$")


def _parse_scalar(s: str) -> Any:
    s = s.strip()
    if s == "":
        return None
    if re.fullmatch(r"-?\d+", s):
        try:
            return int(s)
        except Exception:
            return s
    if re.fullmatch(r"-?\d+\.\d+", s):
        try:
            return float(s)
        except Exception:
            return s
    if s.lower() in ("true", "false"):
        return s.lower() == "true"
    return s


def parse_print_r(text: str) -> Dict[str, Any]:
    """Parse PHP print_r output (stdClass Object / Array nesting) into dict/list."""
    lines = [ln.rstrip("\r") for ln in text.splitlines()]
    root: Dict[str, Any] = {}
    current: Any = root
    stack: List[Any] = []

    def set_in(container: Any, key: str, value: Any) -> None:
        if isinstance(container, list):
            # numeric indexes inside "Array"
            try:
                idx = int(key)
            except Exception:
                container.append({key: value})
                return
            while len(container) <= idx:
                container.append(None)
            container[idx] = value
        else:
            container[key] = value

    for raw_line in lines:
        line = raw_line.strip()
        if not line or line == "...":
            continue
        if line in ("stdClass Object", "Array", "("):
            continue
        if line == ")":
            if stack:
                current = stack.pop()
            continue

        m = _KEYVAL_RE.match(raw_line)
        if not m:
            continue

        key = m.group(1)
        val = m.group(2).strip()

        if val in ("Array", "stdClass Object"):
            new_container: Any = [] if val == "Array" else {}
            set_in(current, key, new_container)
            stack.append(current)
            current = new_container
        else:
            set_in(current, key, _parse_scalar(val))

    return root


def extract_print_r_from_html(html: str) -> str:
    """Extract the RAW output block from the API-Reader page."""
    # Most commonly it is inside <pre>...</pre>. Fallback: find 'stdClass Object'.
    m = re.search(r"<pre[^>]*>([\s\S]*?)</pre>", html, re.IGNORECASE)
    if m:
        # unescape minimal HTML entities
        block = m.group(1)
        block = block.replace("&gt;", ">").replace("&lt;", "<").replace("&amp;", "&")
        return block

    idx = html.find("stdClass Object")
    if idx >= 0:
        return html[idx:]

    raise ValueError("RAW output not found in response")


def fetch_nomor(raw_apiid: str) -> Dict[str, Any]:
    """Fetch RAW output from NoMoreAngel API-Reader and parse it."""
    url = "https://nomoreangel.de/api-reader/"
    params = {
        "apiid": raw_apiid,
        "rawOut": "on",
    }
    r = requests.get(url, params=params, timeout=30)
    r.raise_for_status()
    raw_block = extract_print_r_from_html(r.text)
    return parse_print_r(raw_block)


# -----------------------------
# Summarization / grouping
# -----------------------------
SHIP_TYPE_NAME = {
    # Ships (common)
    202: "Small Cargo",
    203: "Large Cargo",
    204: "Light Fighter",
    205: "Heavy Fighter",
    206: "Cruiser",
    207: "Battleship",
    208: "Colony Ship",
    209: "Recycler",
    210: "Espionage Probe",
    211: "Bomber",
    213: "Destroyer",
    214: "Deathstar",
    215: "Battlecruiser",
    218: "Reaper",
    219: "Pathfinder",
}


def summarize(parsed: Dict[str, Any]) -> Dict[str, Any]:
    g = parsed.get("generic", {}) or {}

    attackers: List[Dict[str, Any]] = [a for a in (parsed.get("attackers") or []) if isinstance(a, dict)]

    # Unique players + first seen tag
    tag_map: Dict[str, str] = {}
    for a in attackers:
        name = a.get("fleet_owner")
        if not name:
            continue
        if name not in tag_map:
            tag_map[name] = a.get("fleet_owner_alliance_tag") or ""

    players = sorted(tag_map.keys())

    return {
        "generic": {
            "cr_id": g.get("cr_id"),
            "event_time": g.get("event_time"),
            "combat_coordinates": g.get("combat_coordinates"),
            "combat_rounds": g.get("combat_rounds"),
            "winner": g.get("winner"),
            "moon_chance": g.get("moon_chance"),
            "loot_percentage": g.get("loot_percentage"),
            "loot_metal": g.get("loot_metal"),
            "loot_crystal": g.get("loot_crystal"),
            "loot_deuterium": g.get("loot_deuterium"),
            "units_lost_attackers": g.get("units_lost_attackers"),
            "units_lost_defenders": g.get("units_lost_defenders"),
            "debris_metal": g.get("debris_metal"),
            "debris_crystal": g.get("debris_crystal"),
            "debris_deuterium": g.get("debris_deuterium"),
            "wreck_metal": g.get("wreckfield_metal"),
            "wreck_crystal": g.get("wreckfield_crystal"),
            "wreck_deuterium": g.get("wreckfield_deuterium"),
        },
        "attackers": [{"name": p, "tag": tag_map.get(p, "")} for p in players],
    }


# -----------------------------
# Spreadsheet generation
# -----------------------------

def build_workbook(summary: Dict[str, Any], weights: Dict[str, float] | None = None) -> io.BytesIO:
    if not os.path.exists(TEMPLATE_XLSX):
        raise FileNotFoundError(
            "template.xlsx not found. Put your template in the app folder as template.xlsx"
        )

    wb = load_workbook(TEMPLATE_XLSX)

    # Create/overwrite the Report sheet
    if "Report" in wb.sheetnames:
        del wb["Report"]
    ws = wb.create_sheet("Report", 0)

    g = summary["generic"]
    attackers = summary["attackers"]
    weights = weights or {a["name"]: 1.0 for a in attackers}

    # Styles
    header_fill = PatternFill("solid", fgColor="1F2937")
    white = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True, color="111827")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="CBD5E1")
    med = Side(style="medium", color="64748B")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column widths
    for col, w in {1: 24, 2: 22, 3: 20, 4: 18, 5: 18, 6: 18, 7: 18, 8: 18}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "OGame – Spartizione Detriti (NoMoreAngel API Reader)"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].fill = header_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    # Metadata
    ws["A2"] = "Dettagli"
    ws["A2"].font = Font(bold=True, color="374151")

    meta = [
        ("CR-ID", g.get("cr_id")),
        ("Data/Ora", g.get("event_time")),
        ("Coordinate", g.get("combat_coordinates")),
        ("Round", g.get("combat_rounds")),
        ("Vincitore", g.get("winner")),
        ("Moon chance", (g.get("moon_chance") or 0) / 100),
        ("Loot %", (g.get("loot_percentage") or 0) / 100),
    ]

    start_row = 3
    for i, (k, v) in enumerate(meta):
        r = start_row + i
        ws[f"A{r}"] = k
        ws[f"A{r}"].font = bold
        ws[f"A{r}"].alignment = left
        ws[f"B{r}"] = v
        ws[f"B{r}"].alignment = left
        ws[f"A{r}"].border = grid
        ws[f"B{r}"].border = grid

    ws["B8"].number_format = "0.0%"  # moon
    ws["B9"].number_format = "0.0%"  # loot

    # Resource summary
    ws["D2"] = "Riepilogo risorse"
    ws["D2"].font = Font(bold=True, color="374151")

    def _num(x: Any) -> Any:
        return x if x is not None else "n/d"

    summary_rows = [
        ("Loot Metal", _num(g.get("loot_metal"))),
        ("Loot Crystal", _num(g.get("loot_crystal"))),
        ("Loot Deuterium", _num(g.get("loot_deuterium"))),
        ("Loot Totale", "=SUM(D3:D5)"),
        ("Perdite Attaccanti", _num(g.get("units_lost_attackers"))),
        ("Perdite Difensore", _num(g.get("units_lost_defenders"))),
        ("Debris Metal", _num(g.get("debris_metal"))),
        ("Debris Crystal", _num(g.get("debris_crystal"))),
        ("Debris Deuterium", _num(g.get("debris_deuterium"))),
        ("Debris Totale", "=SUM(D9:D11)"),
        ("Wreck Metal (Def)", _num(g.get("wreck_metal"))),
        ("Wreck Crystal (Def)", _num(g.get("wreck_crystal"))),
        ("Wreck Deuterium (Def)", _num(g.get("wreck_deuterium"))),
        ("Wreck Totale", "=IF(OR(D13=\"n/d\",D14=\"n/d\",D15=\"n/d\"),\"n/d\",SUM(D13:D15))"),
    ]

    r0 = 3
    for i, (k, v) in enumerate(summary_rows):
        r = r0 + i
        ws[f"C{r}"] = k
        ws[f"C{r}"].font = bold
        ws[f"C{r}"].alignment = left
        ws[f"D{r}"] = v
        ws[f"D{r}"].alignment = right
        ws[f"C{r}"].border = grid
        ws[f"D{r}"].border = grid
        if isinstance(v, (int, float)):
            ws[f"D{r}"].number_format = "#,##0"

    # Attacker split
    ws["A12"] = "Attaccanti (pesi modificabili)"
    ws["A12"].font = Font(bold=True, color="374151")

    table_start = 13
    headers = ["Giocatore", "TAG", "Peso", "% quota", "DF Metal", "DF Crystal", "DF Deut", "DF Totale"]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(table_start, j, h)
        cell.font = white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = grid

    # Write attacker rows
    first_row = table_start + 1
    for i, a in enumerate(attackers):
        r = first_row + i
        name = a["name"]
        ws[f"A{r}"] = name
        ws[f"B{r}"] = a.get("tag", "")
        ws[f"C{r}"] = float(weights.get(name, 1.0))
        ws[f"D{r}"] = f"=C{r}/SUM($C${first_row}:$C${first_row + len(attackers) - 1})"

        # Base values are in D9/D10/D11
        ws[f"E{r}"] = f"=ROUND($D$9*D{r},0)"
        ws[f"F{r}"] = f"=ROUND($D$10*D{r},0)"
        ws[f"G{r}"] = f"=ROUND($D$11*D{r},0)"
        ws[f"H{r}"] = f"=E{r}+F{r}+G{r}"

        for col in "ABCDEFGH":
            ws[f"{col}{r}"].border = grid
            ws[f"{col}{r}"].alignment = right if col in "CDEFGH" else left

        ws[f"D{r}"].number_format = "0.00%"
        for col in "EFGH":
            ws[f"{col}{r}"].number_format = "#,##0"

    # Fix rounding drift: last row becomes remainder
    if attackers:
        last_r = first_row + len(attackers) - 1
        ws[f"E{last_r}"] = f"=$D$9-SUM(E{first_row}:E{last_r-1})"
        ws[f"F{last_r}"] = f"=$D$10-SUM(F{first_row}:F{last_r-1})"
        ws[f"G{last_r}"] = f"=$D$11-SUM(G{first_row}:G{last_r-1})"
        ws[f"H{last_r}"] = f"=E{last_r}+F{last_r}+G{last_r}"

    # Totals row
    tot_r = first_row + len(attackers)
    ws[f"A{tot_r}"] = "Totale"
    ws[f"D{tot_r}"] = f"=SUM(D{first_row}:D{tot_r-1})"
    ws[f"E{tot_r}"] = f"=SUM(E{first_row}:E{tot_r-1})"
    ws[f"F{tot_r}"] = f"=SUM(F{first_row}:F{tot_r-1})"
    ws[f"G{tot_r}"] = f"=SUM(G{first_row}:G{tot_r-1})"
    ws[f"H{tot_r}"] = f"=SUM(H{first_row}:H{tot_r-1})"

    for col in "ABCDEFGH":
        c = ws[f"{col}{tot_r}"]
        c.border = Border(left=med, right=med, top=med, bottom=med)
        c.fill = PatternFill("solid", fgColor="F3F4F6")
        c.alignment = right if col in "CDEFGH" else left

    ws[f"D{tot_r}"].number_format = "0.00%"
    for col in "EFGH":
        ws[f"{col}{tot_r}"].number_format = "#,##0"

    # Note
    note_row = tot_r + 3
    ws.merge_cells(f"A{note_row}:H{note_row+2}")
    ws[f"A{note_row}"] = (
        "Dati letti dal RAW di NoMoreAngel API-Reader. Modifica i pesi in colonna C per cambiare la spartizione."
    )
    ws[f"A{note_row}"].alignment = wrap
    ws[f"A{note_row}"].font = Font(color="374151", size=10)

    ws.freeze_panes = "A13"

    # Save to memory
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# -----------------------------
# Routes
# -----------------------------

def _parse_weights(raw: str) -> Dict[str, float]:
    """Parse 'Name=1, Other=2' into dict."""
    out: Dict[str, float] = {}
    raw = (raw or "").strip()
    if not raw:
        return out
    for part in raw.split(","):
        if "=" not in part:
            continue
        k, v = part.split("=", 1)
        k = k.strip()
        try:
            out[k] = float(v.strip())
        except Exception:
            continue
    return out


@app.get("/")
def index():
    return render_template("index.html")


@app.post("/preview")
def preview():
    apiid = request.form.get("apiid", "").strip()
    weights = _parse_weights(request.form.get("weights", ""))
    parsed = fetch_nomor(apiid)
    summ = summarize(parsed)
    return render_template("report.html", data=summ, weights=weights)


@app.get("/api/json")
def api_json():
    apiid = request.args.get("apiid", "").strip()
    parsed = fetch_nomor(apiid)
    summ = summarize(parsed)
    return jsonify(summ)


@app.get("/download/xlsx")
def download_xlsx():
    apiid = request.args.get("apiid", "").strip()
    weights = _parse_weights(request.args.get("weights", ""))
    parsed = fetch_nomor(apiid)
    summ = summarize(parsed)

    # ensure every player has a weight
    for a in summ["attackers"]:
        weights.setdefault(a["name"], 1.0)

    bio = build_workbook(summ, weights=weights)
    filename = f"Spartizione_Detriti_{apiid}.xlsx".replace(":", "_")
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)
